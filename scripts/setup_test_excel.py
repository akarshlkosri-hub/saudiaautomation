import openpyxl
import os
import shutil

src = r"C:\Users\ashus\PNR\PNR_Data.xlsx"
dest = r"C:\Users\ashus\SaudiaAutomation\data\passengers.xlsx"

# Copy the original file
print(f"Copying {src} to {dest}...")
shutil.copy(src, dest)

# Load the copied file to expand headers and insert test data
wb = openpyxl.load_workbook(dest)
ws = wb.active

# Define new headers
headers = [
    "PNR", "Last Name", "FULL NAME", "GMAIL ID", "FFN", 
    "PASSWORD", "TICKET NUMBER", "CLASS", "SEC 1", "SEC2", 
    "STATUS", "REMARKS"
]

# Set headers in Row 1
for col_idx, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_idx, value=header)

# Insert dummy data for row 2 to test the Excel reader
ws.cell(row=2, column=4, value="client1@gmail.com") # Gmail
ws.cell(row=2, column=5, value="1006210619")       # FFN
ws.cell(row=2, column=6, value="Hunny@1234")       # Password

# Insert dummy data for row 3 (another passenger with same Gmail)
ws.cell(row=3, column=1, value="9ABCDE")
ws.cell(row=3, column=2, value="TESTER")
ws.cell(row=3, column=4, value="client1@gmail.com") # Same Gmail
ws.cell(row=3, column=5, value="2008340722")       # FFN
ws.cell(row=3, column=6, value="Test@5678")       # Password

# Insert dummy data for row 4 (different Gmail)
ws.cell(row=4, column=1, value="8XYZ12")
ws.cell(row=4, column=2, value="DEMO")
ws.cell(row=4, column=4, value="client2@gmail.com") # Different Gmail
ws.cell(row=4, column=5, value="3009450833")       # FFN
ws.cell(row=4, column=6, value="Demo@9012")       # Password

wb.save(dest)
wb.close()
print("passengers.xlsx successfully created and populated with test data.")
