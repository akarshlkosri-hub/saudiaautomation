import openpyxl
import os
from openpyxl.utils import get_column_letter
import config

def get_pending_passengers():
    """
    Reads the Excel file and returns a list of passengers who have no status
    (Status column is empty or None), sorted by Gmail ID to minimize account switching.
    """
    if not os.path.exists(config.EXCEL_PATH):
        raise FileNotFoundError(f"Excel file not found at: {config.EXCEL_PATH}")
        
    wb = openpyxl.load_workbook(config.EXCEL_PATH)
    ws = wb.active
    
    passengers = []
    
    # Data starts from row 2 (row 1 is header)
    for r in range(2, ws.max_row + 1):
        ffn_val = ws.cell(row=r, column=config.COL_FFN).value
        pass_val = ws.cell(row=r, column=config.COL_PASSWORD).value
        status_val = ws.cell(row=r, column=config.COL_STATUS).value
        gmail_val = ws.cell(row=r, column=config.COL_GMAIL).value
        pnr_val = ws.cell(row=r, column=config.COL_PNR).value
        
        # If Status already exists (like 'Success' or 'Failed'), skip it
        if status_val is not None and str(status_val).strip() != "":
            continue
            
        # FFN and Password are required to log in
        if not ffn_val or not pass_val:
            continue
            
        passengers.append({
            'row': r,
            'pnr': str(pnr_val).strip() if pnr_val else "",
            'gmail': str(gmail_val).strip() if gmail_val else "",
            'ffn': str(ffn_val).strip(),
            'password': str(pass_val).strip()
        })
        
    wb.close()
    
    # Sort by Gmail ID to group same-Gmail accounts together
    # Items with empty/no gmail go last
    passengers.sort(key=lambda x: x['gmail'] if x['gmail'] else "zzzzzzzz")
    
    return passengers

def update_status(row_num, status, remarks=""):
    """
    Updates the Status and Remarks columns for a specific row in the Excel sheet.
    """
    if not os.path.exists(config.EXCEL_PATH):
        raise FileNotFoundError(f"Excel file not found at: {config.EXCEL_PATH}")
        
    wb = openpyxl.load_workbook(config.EXCEL_PATH)
    ws = wb.active
    
    ws.cell(row=row_num, column=config.COL_STATUS, value=status)
    ws.cell(row=row_num, column=config.COL_REMARKS, value=remarks)
    
    wb.save(config.EXCEL_PATH)
    wb.close()

def get_gmail_for_row(row_num):
    """
    Gets the Gmail address stored in Column D for a given row.
    """
    if not os.path.exists(config.EXCEL_PATH):
        raise FileNotFoundError(f"Excel file not found at: {config.EXCEL_PATH}")
        
    wb = openpyxl.load_workbook(config.EXCEL_PATH)
    ws = wb.active
    
    gmail_val = ws.cell(row=row_num, column=config.COL_GMAIL).value
    wb.close()
    
    return str(gmail_val).strip() if gmail_val else ""

if __name__ == "__main__":
    # Test script if executed directly
    print("Testing excel_reader.py...")
    try:
        passengers = get_pending_passengers()
        print(f"Successfully read {len(passengers)} pending passengers:")
        for p in passengers:
            print(f"  Row {p['row']}: FFN={p['ffn']} | Gmail={p['gmail']}")
            
        if passengers:
            test_row = passengers[0]['row']
            print(f"Writing test status to row {test_row}...")
            update_status(test_row, "PendingTest", "Temporary test remark")
            print("Status updated successfully.")
            
            # Read again to verify
            wb = openpyxl.load_workbook(config.EXCEL_PATH)
            ws = wb.active
            status = ws.cell(row=test_row, column=config.COL_STATUS).value
            remarks = ws.cell(row=test_row, column=config.COL_REMARKS).value
            print(f"Verification: Status={status}, Remarks={remarks}")
            
            # Revert the test status
            update_status(test_row, None, None)
            print("Test status reverted.")
    except Exception as e:
        print(f"Error during test: {e}")
